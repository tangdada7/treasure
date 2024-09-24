#!/usr/bin/env python# -*- coding: utf-8 -*-# @Time : 2024/9/18 下午4:37# @Author : MiuziZhouimport requestsfrom bs4 import BeautifulSoup  # 用来获取网页内容
import re
import openpyxl
from math import radians, cos, sin, asin, sqrt
import pandas as pd
import json
from urllib.request import urlopen
from urllib.parse import quote
import traceback
import time

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36 Edg/128.0.0.0',
    'Host': 'gz.lianjia.com'
}

'''第一步：爬取小区信息'''


def export_communityInfo(xiaoquInfo_list):
    '''导出小区信息'''
    with open('广州地区小区信息.txt', 'a', encoding='utf-8') as file:
        file.write(','.join(xiaoquInfo_list))
        file.write('\n')


def get_all_houses_info(houses_url, community_name):
    '''从小区的所有在售房屋链接获取每个房屋的信息并导出'''
    r = requests.get(url=houses_url, headers=headers)
    soup = BeautifulSoup(r.text, 'lxml')
    houses = soup.find_all('div', class_='info clear')

    for house in houses:
        try:
            price = house.find('div', class_='totalPrice').span.text + '万'
            area = house.find('div', class_='houseInfo').text.split('|')[1].strip()
            export_house_info([community_name, price, area])  # 导出房屋信息到文件
        except Exception as e:
            print(f'房屋信息提取出错: {e}')


def export_house_info(house_info):
    '''导出房屋信息到文件'''
    with open('小区房屋信息.txt', 'a', encoding='utf-8') as file:
        file.write(','.join(house_info))
        file.write('\n')


# community_url是小区的详细网页
def get_communityInfo(community_url, community_name):
    '''获取某个小区的信息'''
    r = requests.get(url=community_url, headers=headers)
    soup = BeautifulSoup(r.text, 'lxml')

    # 尝试获取指向小区所有在售二手房的链接
    try:
        all_houses_link = soup.find('a', class_='fr', text='查看小区全部在售二手房')['href']
        get_all_houses_info(all_houses_link, community_name)  # 获取所有房屋信息的函数调用
    except Exception as e:
        print(f'无法获取 {community_name} 的在售房屋链接: {e}')


    # 尝试获取小区均价
    try:
        unitPrice = soup.find(name='span', attrs={'class': 'xiaoquUnitPrice'}).text  # 小区均价
    except:
        unitPrice = '空'


    # 尝试获取小区地址
    try:
        address = soup.find(name='div', attrs={'class': 'detailDesc'}).text  # 小区地址
        address = '"' + address + '"'
    except:
        address = '空'

    # 获取其他小区信息
    xiaoquInfo = soup.find_all(name='span', attrs={'class': 'xiaoquInfoContent'})  # 小区信息
    xiaoquInfo_list = []

    # 将小区名称、地址和均价添加到列表
    community_name = '"' + community_name + '"'
    xiaoquInfo_list.append(community_name)
    xiaoquInfo_list.append(address)
    xiaoquInfo_list.append(unitPrice)

    # 遍历每个小区信息，并处理 NoneType 问题
    for info in xiaoquInfo:
        # 如果 info.string 是 None，替换为 '空'
        if info.string is None:
            xiaoquInfo_list.append('空')
        else:
            xiaoquInfo_list.append(info.string)

    # 移除最后一个冗余的元素（如果有必要）
    if xiaoquInfo_list:
        xiaoquInfo_list.pop()

    # 导出小区信息到文件
    export_communityInfo(xiaoquInfo_list)

    print(f'已爬取 {community_name} 的信息')


def xiaoqu_pachong():
    '''获取所有小区名字和链接'''
    for i in range(1, 100):
        page_num = "" if i == 1 else 'pg' + str(i)
        url = r"https://gz.lianjia.com/xiaoqu/" + page_num + "rs广州/"
        r = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(r.text, 'lxml')
        for j in soup.find_all(name='div', attrs={'class': 'title'}):
            community = str(j)
            if '''target="_blank"''' in community:
                community_list = re.search(r'<a href="(.*?)" target="_blank">(.*?)</a>', community)
                community_url = community_list.group(1)
                community_name = community_list.group(2)
                try:
                    get_communityInfo(community_url, community_name)
                except Exception as e:
                    # 打印详细的错误信息
                    print(f'第{i}页的{community_name}出错了, url是{community_url}，错误信息：{e}')
                    traceback.print_exc()

        # 添加随机延迟，避免反爬虫
        time.sleep(2)


'''第二步：数据清洗'''


# # 处理建成年份的函数
# def process_year(year):
#     if isinstance(year, str):
#         year = year.strip('年 ').strip()  # 移除多余的字符
#         # 检查是否是年份区间
#         if '-' in year:
#             start_year, end_year = year.split('-')
#             try:
#                 return str(int((int(start_year) + int(end_year)) / 2))  # 返回中间年份
#             except ValueError:
#                 return None
#         try:
#             return str(int(re.sub(r'\D', '', year)))  # 只保留数字部分
#         except ValueError:
#             return None
#     return None
#
#
# # 数据清洗函数
# def clean_data():
#     '''数据清洗'''
#     # 读取原始数据
#     column_name = ['小区名称', '小区地址', '均价', '建筑类型', '住户数', '楼栋数', '容积率', '绿化率', '物业类型',
#                    '建成年份', '供暖信息', '水电', '物业费', '门店信息', '物业公司']
#     try:
#         xiaoqu_data = pd.read_csv('广州地区小区信息.txt', names=column_name, encoding='utf-8')
#         print("数据读取成功，数据行数: ", len(xiaoqu_data))
#     except FileNotFoundError:
#         print("文件未找到，请确认路径是否正确。")
#         return
#
#     # 处理建成年份
#     xiaoqu_data['建成年份'] = xiaoqu_data['建成年份'].apply(lambda x: process_year(x) if isinstance(x, str) else x)
#     print("处理建成年份后：")
#     print(xiaoqu_data[['小区名称', '建成年份']].head())
#
#     # 移除楼栋数和住户数中的非数字字符，并处理百分比
#     xiaoqu_data['楼栋数'] = xiaoqu_data['楼栋数'].str.extract(r'(\d+)').astype(float)  # 提取数字部分
#     xiaoqu_data['住户数'] = xiaoqu_data['住户数'].str.extract(r'(\d+)').astype(float)  # 提取数字部分
#
#     # 查看楼栋数和住户数处理后的结果
#     print("处理楼栋数和住户数后：")
#     print(xiaoqu_data[['小区名称', '楼栋数', '住户数']].head())
#
#     # 删除 "暂无信息" 和空值
#     xiaoqu_data = xiaoqu_data.dropna(subset=['建成年份', '楼栋数', '住户数', '均价', '小区地址'])
#
#     print("清除'暂无信息'和空值后的数据行数：", len(xiaoqu_data))
#
#     # 转换建成年份为整数
#     xiaoqu_data['建成年份'] = pd.to_numeric(xiaoqu_data['建成年份'], errors='coerce')
#
#     # 过滤建成年份 >= 1975 和住户数 >= 10 的数据
#     xiaoqu_data = xiaoqu_data[(xiaoqu_data['建成年份'] >= 1975) & (xiaoqu_data['住户数'] >= 10)]
#     print("过滤建成年份 >= 1975 和住户数 >= 10 后的数据行数：", len(xiaoqu_data))
#
#     # 添加新的字段：楼均住户数和小区年龄
#     xiaoqu_data['楼均住户数'] = xiaoqu_data['住户数'] / xiaoqu_data['楼栋数']
#     xiaoqu_data['小区年龄'] = 2020 - xiaoqu_data['建成年份']
#
#     # 查看新增字段
#     print("新增字段楼均住户数和小区年龄后：")
#     print(xiaoqu_data[['小区名称', '楼均住户数', '小区年龄']].head())
#
#     # 分离区域和地址
#     xiaoqu_address = xiaoqu_data['小区地址'].str.split(')', expand=True, n=1)
#
#     if xiaoqu_address.shape[1] == 2:
#         xiaoqu_address.columns = ['区域', '地址']
#         xiaoqu_address['区域'] = xiaoqu_address['区域'].str[1:]
#         xiaoqu_address['地址'] = '广州市' + xiaoqu_address['地址']
#     else:
#         xiaoqu_address = pd.DataFrame({'区域': '未知区域', '地址': xiaoqu_data['小区地址']})
#
#     xiaoqu_data = pd.merge(xiaoqu_data, xiaoqu_address, how='left', left_index=True, right_index=True)
#
#     print("地址信息分离和合并后：")
#     print(xiaoqu_data[['小区名称', '区域', '地址']].head())
#
#     # 过滤楼均住户数 <= 200 的数据
#     xiaoqu_data = xiaoqu_data[xiaoqu_data['楼均住户数'] <= 200]
#
#     # 删除重复的小区
#     xiaoqu_data.drop_duplicates('小区名称', inplace=True)
#
#     print("删除重复小区后的数据行数：", len(xiaoqu_data))
#
#     # 输出到Excel
#     try:
#         xiaoqu_data.to_excel(r"广州小区数据_clean.xlsx", index=False, engine='openpyxl')
#         print("数据清洗完成并成功输出到 Excel 文件中。")
#     except Exception as e:
#         print(f"导出 Excel 文件失败: {e}")


def clean_export_excel_community(input_filename, output_filename):
    # 读取文本文件，假设字段是通过逗号分隔的
    try:
        data = pd.read_csv(input_filename, header=None, sep=',', encoding='utf-8')
        print("原始数据列数:", data.shape[1])  # 打印列数检查
    except Exception as e:
        print(f"读取文件时出错：{e}")
        return

        # 列名（根据实际列数调整）
    column_names = ['小区名称', '地址', '小区均价', '建筑类型', '房屋总数', '楼栋总数', '绿化率', '容积率', '交易权属',
                    '建成年代', '供暖类型', '用水类型', '用电类型', '物业费', '附近门店信息', '物业公司']
    if len(column_names) >= data.shape[1]:
        data.columns = column_names[:data.shape[1]]  # 只使用前n个列名
    else:
        print("警告: 列名数量少于数据列数量，将使用数字索引作为额外的列名。")
        data.columns = column_names + [f"额外信息{i}" for i in range(data.shape[1] - len(column_names))]

    # 数据清洗，例如处理缺失数据、转换数据类型等
    data.replace({'空': pd.NA}, inplace=True)

    # 去重，保留小区名称的第一条记录
    data.drop_duplicates(subset=['小区名称'], keep='first', inplace=True)

    # 导出到Excel
    try:
        data.to_excel(output_filename, index=False)
        print(f"数据已成功导出到 {output_filename}")
    except Exception as e:
        print(f"导出Excel文件时出错：{e}")


def export_excel_house(input_filename, output_filename):
    # 读取数据
    try:
        data = pd.read_csv(input_filename, header=None, sep=',', names=['小区名称', '总价', '平方米'], encoding='utf-8')
        # 导出数据到Excel
        data.to_excel(output_filename, index=False, engine='openpyxl')
    except Exception as e:
        print(f"读取或写入文件出错: {e}")
        return

    # 打开Excel文件合并单元格
    try:
        workbook = load_workbook(output_filename)
        sheet = workbook.active

        # 居中对齐设置
        align_center = Alignment(horizontal='center', vertical='center')

        # 获取小区名称列
        column = sheet['A']
        previous_value = None
        start_index = 1
        # 从第二行开始遍历（跳过标题）
        for index, cell in enumerate(column[1:], start=2):
            # 设置单元格居中
            sheet.cell(row=index, column=1).alignment = align_center
            sheet.cell(row=index, column=2).alignment = align_center
            sheet.cell(row=index, column=3).alignment = align_center
            if cell.value != previous_value:
                if index - start_index > 1:
                    sheet.merge_cells(start_row=start_index, start_column=1, end_row=index - 1, end_column=1)
                start_index = index
            previous_value = cell.value
        # 处理最后一个合并区域
        if index - start_index >= 1:
            sheet.merge_cells(start_row=start_index, start_column=1, end_row=index, end_column=1)

        workbook.save(output_filename)
        workbook.close()
        print("\n小区房屋信息处理完毕，并已保存。")
    except Exception as e:
        print(f"处理Excel文件时出错: {e}")


'''第三步：获取经纬度和到人民广场的距离'''


def get_bd_distance():
    '''获得地址的经纬度和到人民广场的距离（百度）'''
    wb = openpyxl.load_workbook('广州小区数据.xlsx')
    ws = wb['Sheet1']
    maxrow = ws.max_row
    for i in range(2, maxrow + 1):
        address = ws["C" + str(i)].value.split(',')[0]
        address2 = '广州市' + ws["A" + str(i)].value
        try:
            temp_list = getjwd_bd(address)
        except:
            try:
                temp_list = getjwd_bd(address2)
            except:
                temp_list = getjwd_bd(address)
        lng = temp_list[0]
        lat = temp_list[1]
        bd_rmgc_lat = 23.12908
        bd_rmgc_lng = 113.26436
        distance = get_distance(lng, lat, bd_rmgc_lng, bd_rmgc_lat)
        ws['J1'].value = '到人民广场的距离（百度）'
        ws["J" + str(i)].value = distance
        print("{}的经度是{}，纬度是{}，到人民广场的距离是{}".format(address, lng, lat, distance))
    wb.save('广州小区数据.xlsx')
    wb.close()


def get_gd_distance():
    '''获得地址的经纬度和到人民广场的距离（高德）'''
    wb = openpyxl.load_workbook('广州小区数据.xlsx')
    ws = wb['Sheet1']
    maxrow = ws.max_row
    for i in range(2, maxrow + 1):
        address = ws["C" + str(i)].value.split(',')[0]
        address2 = '广州市' + ws["A" + str(i)].value
        try:
            temp_list = getjwd_gd(address)
        except:
            try:
                temp_list = getjwd_gd(address2)
            except:
                temp_list = getjwd_gd(address)
        lng = temp_list[0]
        lat = temp_list[1]
        gd_rmgc_lat = 23.12908
        gd_rmgc_lng = 113.26436
        distance = get_distance(lng, lat, gd_rmgc_lng, gd_rmgc_lat)
        ws['K1'].value = '到人民广场的距离（高德）'
        ws["K" + str(i)].value = distance
        print("{}的经度是{}，纬度是{}，到人民广场的距离是{}".format(address, lng, lat, distance))
    wb.save('广州小区数据.xlsx')
    wb.close()


def getjwd_bd(address):
    '''通过百度API获取经纬度'''
    url = 'http://api.map.baidu.com/geocoding/v3/?address={}&output=json&ak=百度API密钥'.format(quote(address))
    temp = urlopen(url)
    temp = json.loads(temp.read())
    lng = temp['result']['location']['lng']
    lat = temp['result']['location']['lat']
    return lng, lat


def getjwd_gd(address):
    '''通过高德API获取经纬度'''
    url = 'https://restapi.amap.com/v3/geocode/geo?key=高德API密钥&address={}'.format(quote(address))
    temp = urlopen(url)
    temp = json.loads(temp.read())
    lng = temp['geocodes'][0]['location'].split(',')[0]
    lat = temp['geocodes'][0]['location'].split(',')[1]
    return lng, lat


def get_distance(lng1, lat1, lng2, lat2):
    '''两点间的距离计算'''
    lng1, lat1, lng2, lat2 = map(radians, [float(lng1), float(lat1), float(lng2), float(lat2)])
    dlon = lng2 - lng1
    dlat = lat2 - lat1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    c = 2 * asin(sqrt(a))
    r = 6371  # 地球半径，单位为公里
    return c * r


if __name__ == '__main__':
    # 爬取广州链家网的小区数据
    # xiaoqu_pachong()

    # 清洗数据
    # clean_data()
    clean_export_excel_community('广州地区小区信息.txt', '广州地区小区信息.xlsx')
    export_excel_house('小区房屋信息.txt','小区房屋信息.xlsx')

# 获取百度经纬度并计算距离
# get_bd_distance()

# 获取高德经纬度并计算距离
# get_gd_distance()
